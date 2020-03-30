import requests
from bs4 import BeautifulSoup
import openpyxl


class ParsingShop:

    def __init__(self):
        self.url = 'https://ekipirovka.ru/catalog/'
        self.catalogUrl = []
        self.infoProducts = {}
        self.info = []

    def pageLoading(self):
        """Создание списка элементов ссылок"""
        print('Загружается страница {}...'.format(self.url))
        res = requests.get(self.url)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, features="html.parser")
        self.catalogElem = soup.select('div.catalog-section-childs a')
        return self.catalogElem

    def createLincs(self, catalogElem):
        """Создание списка ссылок на разделы с товарами"""
        print('Создаётся каталог ссылок...')
        for i in catalogElem:
            address = 'https://ekipirovka.ru' + i.get('href')
            self.catalogUrl.append(address)
        print('Каталог ссылок создан')
        return self.catalogUrl

    def parsingProducts(self, catalogUrl):
        """Создание списка со словарями, включающими данные о товарах"""
        print('Перебор ссылок...')
        for url in catalogUrl:
            while not url.endswith('#'):
                try:
                    # Загрузка страницы
                    print('Загружается страница {}...'.format(url))
                    res = requests.get(url)
                    res.raise_for_status()
                    soup = BeautifulSoup(res.text, features="html.parser")
                    titleElem = soup.select('div.catalog-item-info')
                    for i in titleElem:
                        title = i.find("span", {"itemprop": "name"})
                        title = title.getText().split()
                        title = ' '.join(title)
                        price = i.find("span", {"class": "catalog-item-price"})
                        try:
                            price = price.getText().split()
                            price = ' '.join(price)
                        except:
                            price = 'Нет в наличии'
                        self.infoProducts['title'] = title
                        self.infoProducts['price'] = price
                        self.info.append(self.infoProducts)
                        self.infoProducts = {}
                        print('{} {}'.format(len(self.info), self.info[-1]))

                    prevLink = soup.select('a#navigation_1_next_page')[0]
                    url = 'https://ekipirovka.ru' + prevLink.get('href')

                except Exception as exc:
                    print('Возможно последняя страница: {} адреса {}'.format(exc, url))
                    break

        return self.info


class CreateFile:

    def __init__(self):
        """Создание таблицы"""
        print('Создаётся таблица...')
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 70
        sheet.column_dimensions['B'].width = 30
        sheet.merge_cells('A1:B1')
        sheet.freeze_panes = 'A3'
        sheet['A1'] = 'Прайс-лист спортивного магазина'
        sheet['A2'] = 'Название'
        sheet['B2'] = 'Цена'
        sheet.title = 'Наименование и цена'
        wb.save('sportShop.xlsx')
        print('Таблица создана')

    @staticmethod
    def addContent(info):
        """Занесение данных в таблицу"""
        wb = openpyxl.load_workbook('sportShop.xlsx')
        sheet = wb.active
        print('Занесение данных в таблицу...')
        for i in range(len(info)):
            data = info[i]
            sheet['A' + '{}'.format(str(i + 3))] = data['title']
            sheet['B' + '{}'.format(str(i + 3))] = data['price']
        wb.save('sportShop.xlsx')


